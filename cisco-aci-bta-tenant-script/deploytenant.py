
# Import General Classes
from openpyxl import load_workbook
import urllib3
import argparse

# Import Cobra Access Classes
from cobra.mit.access import MoDirectory
from cobra.mit.request import ConfigRequest
from cobra.mit.session import LoginSession

# Import Cobra Model Classes
from cobra.model.fv import (Tenant, Ctx, BD, RsCtx, Ap, AEPg, RsBd, Subnet, RsDomAtt)
from cobra.model.fvns import (VlanInstP, EncapBlk)
from cobra.model.infra import (Infra, RsVlanNs, AttEntityP, RsDomP, Generic, RsFuncToEpg, FuncP)
from cobra.model.pol import Uni
from cobra.model.phys import DomP

# Parse command line arguments and set to variables
parser = argparse.ArgumentParser()

group = parser.add_argument_group('Required')
group.add_argument("-a", type=str, dest="apic", help="APIC IP/FQDN")
group.add_argument("-u", type=str, dest="user", help="Username")
group.add_argument("-p", type=str, dest="passwd", help="Password")

args = parser.parse_args()

apicUrl = 'https://%s' % (args.apic)
user = args.user
password = args.passwd

DATA = 'spreadsheet.xlsx'

# Disable urllib3 (HTTPS Self-Signed Certificate) warnings
urllib3.disable_warnings()

# Log into the APIC and create directory object
print('Initializing connection to APIC...')
ls = LoginSession(apicUrl, user, password)
md = MoDirectory(ls)
md.login()


def safe_str(obj):
    return str(obj).strip().encode('ascii', errors='ignore').decode()


wb = load_workbook(DATA)
print("Workbook Loaded.")

for sheet in wb.sheetnames:
    ws = wb[sheet]
    max_row = ws.max_row
    for i in range(2, max_row + 1):
        # Create variables by importing values from spreadsheet
        ss_tenant = safe_str(ws.cell(row=i, column=1).value)
        ss_vrf = safe_str(ws.cell(row=i, column=2).value)
        ss_domain = safe_str(ws.cell(row=i, column=3).value)
        ss_vlan_pool = safe_str(ws.cell(row=i, column=4).value)
        ss_aep = safe_str(ws.cell(row=i, column=5).value)
        ss_ap = safe_str(ws.cell(row=i, column=6).value)
        ss_epg = safe_str(ws.cell(row=i, column=7).value)
        ss_epg_alias = safe_str(ws.cell(row=i, column=8).value)
        ss_encap = safe_str(ws.cell(row=i, column=9).value)
        ss_bd = safe_str(ws.cell(row=i, column=10).value)
        ss_bd_alias = safe_str(ws.cell(row=i, column=11).value)
        ss_bd_settings = safe_str(ws.cell(row=i, column=12).value)
        # ss_bd_vmac = safe_str(ws.cell(row=i, column=13).value)
        ss_bd_vmac = 'None'
        # ss_bd_gw = safe_str(ws.cell(row=i, column=14).value)
        ss_bd_gw = 'None'

        # Needed Object DN's
        vlan_pool = 'uni/infra/vlanns-[%s]-static' % (ss_vlan_pool)
        domain = 'uni/phys-%s' % (ss_domain)
        tenant = 'uni/tn-%s/ap-%s/epg-%s' % (ss_tenant, ss_ap, ss_epg)

        # Define top level Policy Universe
        polUni = Uni('')

        # Create Domain and VLAN Pool
        uniInfra = Infra(polUni)
        fvnsVlanInstP = VlanInstP(uniInfra, name=ss_vlan_pool, allocMode='static')
        fvEncapBlk = EncapBlk(fvnsVlanInstP, from_=ss_encap, to=ss_encap, role='external', allocMode='inherit')
        fvDomain = DomP(polUni, name=ss_domain)
        infraRsVlanNs = RsVlanNs(fvDomain, tDn=vlan_pool)

        # Create Tenant
        fvTenant = Tenant(polUni, ss_tenant)

        # Create VRF
        fvCtx = Ctx(fvTenant, name=ss_vrf)

        # Create ANP
        fvAp = Ap(fvTenant, name=ss_ap)

        # Create EPG
        fvAEPg = AEPg(fvAp, name=ss_epg, nameAlias=ss_epg_alias)
        # Associate Physical Domain with EPG
        fvRsDomAtt = RsDomAtt(fvAEPg, tDn=domain)

        # Create Bridge Domain
        # - looks for forward slash separated key=value pairs to determine bd_settings
        # - looks at ss_bd_gw to determine is subnet is needed
        # - looks at ss_bd_vmac to determine if virtual mac is needed
        bd_arpflood = ''
        bd_unicastroute = ''
        bd_unkucast = ''

        for pair in ss_bd_settings.split('/'):
            if pair.split(':')[0].strip() == 'UniRtg':
                bd_unicastroute = pair.split(':')[1].lower().strip()
            elif pair.split(':')[0].strip() == 'L2UnkUni':
                bd_unkucast = pair.split(':')[1].lower().strip()
            elif pair.split(':')[0].strip() == 'ArpFlood':
                bd_arpflood = pair.split(':')[1].lower().strip()

        if ss_bd_vmac != 'None':
            fvBD = BD(fvTenant, name=ss_bd, vmac=ss_bd_vmac, limitIpLearnToSubnets='yes', mcastAllow='no', unkMcastAct='flood', unkMacUcastAct=bd_unkucast, arpFlood=bd_arpflood, unicastRoute=bd_unicastroute,
                      multiDstPktAct='bd-flood', type='regular', ipLearning='yes', nameAlias=ss_bd_alias)
            if ss_bd_gw != 'None':
                fvSubnet = Subnet(fvBD, ip=ss_bd_gw, virtual='yes', scope='public', preferred='yes', ctrl='')
        else:
            fvBD = BD(fvTenant, name=ss_bd, vmac='not-applicable', limitIpLearnToSubnets='yes', mcastAllow='no', unkMcastAct='flood', unkMacUcastAct=bd_unkucast, arpFlood=bd_arpflood, unicastRoute=bd_unicastroute,
                      multiDstPktAct='bd-flood', type='regular', ipLearning='yes', nameAlias=ss_bd_alias)
            if ss_bd_gw != 'None':
                fvSubnet = Subnet(fvBD, ip=ss_bd_gw, virtual='no', scope='public', preferred='yes', ctrl='')

        # Associate Bridge Domain with VRF
        fvRsCtx = RsCtx(fvBD, tnFvCtxName=ss_vrf)

        # Associate EPG with Bridge Domains
        fvRsBd = RsBd(fvAEPg, tnFvBDName=ss_bd)

        # Create AEP and Associate EPG with AEPs
        # - Looks for comma separated values in xlsx cell
        for aep_item in ss_aep.split(','):
            infraAttEntityP = AttEntityP(uniInfra, ownerKey='', name=aep_item, desc='', ownerTag='', nameAlias='', annotation='')
            infraGeneric = Generic(infraAttEntityP, nameAlias='', annotation='', descr='', name='default')
            infraRsFuncToEpg = RsFuncToEpg(infraGeneric, tDn=tenant, primaryEncap='unknown', instrImedcy='lazy', mode='regular', encap=ss_encap, annotation='')
            infraRsDomP = RsDomP(infraAttEntityP, annotation='', tDn=domain)
            infraFuncP = FuncP(uniInfra)

        # Commit to APIC
        c = ConfigRequest()
        c.addMo(fvTenant)
        c.addMo(fvnsVlanInstP)
        c.addMo(fvDomain)
        c.addMo(uniInfra)
        md.commit(c)

    print("Tab: " + sheet + " ** COMPLETED **")
